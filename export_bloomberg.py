"""
Bloomberg Excel → JSON Export Script

Usage (on Bloomberg terminal or any machine with the Excel file):
    python export_bloomberg.py bloomberg_KO.xlsx

Reads the populated Bloomberg Excel template and exports a structured JSON file
that the equity analysis app can consume.

Output: data/{TICKER}.json
"""

import sys
import json
import os
from datetime import datetime

try:
    import openpyxl
except ImportError:
    print("Installing openpyxl...")
    os.system(f"{sys.executable} -m pip install openpyxl")
    import openpyxl


def safe_float(val):
    """Convert a value to float, returning None if not possible."""
    if val is None:
        return None
    try:
        f = float(val)
        return f if f == f else None  # NaN check
    except (ValueError, TypeError):
        return None


def safe_str(val):
    """Convert a value to string, returning None if empty."""
    if val is None:
        return None
    s = str(val).strip()
    if s in ("", "#N/A", "#N/A N/A", "N/A", "#REF!", "#VALUE!", "#DIV/0!"):
        return None
    return s


def export_bloomberg(xlsx_path):
    """Read populated Bloomberg Excel template and export to JSON."""

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    # ─── Profile Sheet ───
    ws = wb["Profile"]
    ticker = safe_str(ws["B1"].value)
    if not ticker:
        print("ERROR: No ticker found in Profile!B1")
        sys.exit(1)

    ticker = ticker.upper().strip()
    exported_at = datetime.now().isoformat()

    # Read profile fields (column A = label, column B = value, column C = json_key)
    profile = {}
    valuation = {}
    consensus = {}
    profitability = {}
    dividends_snapshot = {}
    other = {}

    current_section = None
    for row in ws.iter_rows(min_row=5, max_col=3, values_only=False):
        cell_a = row[0].value
        cell_b = row[1].value
        cell_c = row[2].value  # json_key

        if cell_a and "COMPANY PROFILE" in str(cell_a):
            current_section = "profile"
            continue
        elif cell_a and "VALUATION" in str(cell_a):
            current_section = "valuation"
            continue
        elif cell_a and "ANALYST CONSENSUS" in str(cell_a):
            current_section = "consensus"
            continue
        elif cell_a and "PROFITABILITY" in str(cell_a):
            current_section = "profitability"
            continue
        elif cell_a and "DIVIDENDS" in str(cell_a):
            current_section = "dividends_snapshot"
            continue
        elif cell_a and "OTHER" in str(cell_a):
            current_section = "other"
            continue

        json_key = safe_str(cell_c)
        if not json_key or not current_section:
            continue

        # Determine if numeric or string
        val = cell_b
        if json_key in ("name", "sector", "industry", "sub_industry", "country",
                         "exchange", "currency", "description"):
            parsed = safe_str(val)
        else:
            parsed = safe_float(val)

        target = {
            "profile": profile,
            "valuation": valuation,
            "consensus": consensus,
            "profitability": profitability,
            "dividends_snapshot": dividends_snapshot,
            "other": other,
        }[current_section]

        target[json_key] = parsed

    # ─── Financials Sheet ───
    ws2 = wb["Financials"]
    financials = {}

    # Read year headers (row 4, columns B-F)
    # Read line items (rows 5+)
    for row in ws2.iter_rows(min_row=5, max_col=7, values_only=False):
        label = safe_str(row[0].value)
        json_key = safe_str(row[6].value) if row[6].value else None

        if not label or not json_key:
            continue

        for yr_idx in range(5):
            yr_key = f"FY{yr_idx - 4}" if yr_idx < 4 else "FY0"
            if yr_key not in financials:
                financials[yr_key] = {}

            val = safe_float(row[yr_idx + 1].value)
            financials[yr_key][json_key] = val

    # ─── Peers Sheet ───
    ws3 = wb["Peers"]
    peers = []

    # Read BDS peer list (A5:A14)
    peer_tickers = []
    for r in range(5, 15):
        val = safe_str(ws3.cell(row=r, column=1).value)
        if val:
            # BDS might return "PEP US Equity" or just "PEP"
            val = val.replace(" US Equity", "").strip()
            peer_tickers.append(val)

    # Read peer comparison table (rows 22-32)
    # Row 21 = headers, Row 22 = target, Rows 23-32 = peers
    peer_metrics_keys = [
        "ticker", "name", "price", "market_cap", "pe_ratio", "forward_pe",
        "ev_ebitda", "price_to_sales", "price_to_book", "peg_ratio",
        "gross_margin", "operating_margin", "net_margin", "roe",
        "dividend_yield", "beta", "target_price"
    ]

    peer_data = []
    for r in range(22, 33):  # Target + 10 peers
        row_data = {}
        ticker_val = safe_str(ws3.cell(row=r, column=1).value)
        if not ticker_val:
            continue

        ticker_val = ticker_val.replace(" US Equity", "").strip()
        row_data["ticker"] = ticker_val

        for j, key in enumerate(peer_metrics_keys):
            if j == 0:
                continue  # already got ticker
            val = ws3.cell(row=r, column=j+1).value
            if key == "name":
                row_data[key] = safe_str(val)
            else:
                row_data[key] = safe_float(val)

        row_data["is_target"] = (r == 22)
        peer_data.append(row_data)

    # Read median row
    medians = {}
    for j, key in enumerate(peer_metrics_keys):
        if j < 2:
            continue  # skip ticker and name
        val = safe_float(ws3.cell(row=34, column=j+1).value)
        medians[key] = val

    # ─── Dividends Sheet ───
    ws4 = wb["Dividends"]
    dividend_history = []

    for r in range(5, 15):
        year = ws4.cell(row=r, column=1).value
        dps = safe_float(ws4.cell(row=r, column=2).value)
        growth = safe_float(ws4.cell(row=r, column=3).value)

        if year and dps is not None:
            dividend_history.append({
                "year": int(year) if year else None,
                "dps": dps,
                "growth": growth,
            })

    # ─── Assemble JSON ───
    output = {
        "ticker": ticker,
        "source": "bloomberg",
        "exported_at": exported_at,
        "profile": profile,
        "valuation": valuation,
        "consensus": consensus,
        "profitability": profitability,
        "dividends_snapshot": dividends_snapshot,
        "other": other,
        "financials": financials,
        "peers": peer_tickers,
        "peer_data": peer_data,
        "peer_medians": medians,
        "dividend_history": dividend_history,
    }

    # ─── Write JSON ───
    os.makedirs("data", exist_ok=True)
    output_path = os.path.join("data", f"{ticker}.json")
    with open(output_path, "w") as f:
        json.dump(output, f, indent=2, default=str)

    print(f"Exported {ticker} → {output_path}")
    print(f"  Profile fields: {len(profile)}")
    print(f"  Valuation fields: {len(valuation)}")
    print(f"  Consensus fields: {len(consensus)}")
    print(f"  Financial years: {len(financials)}")
    print(f"  Peers found: {len(peer_tickers)}")
    print(f"  Dividend years: {len(dividend_history)}")

    return output_path


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python export_bloomberg.py <bloomberg_TICKER.xlsx>")
        print("Example: python export_bloomberg.py bloomberg_KO.xlsx")
        sys.exit(1)

    export_bloomberg(sys.argv[1])
